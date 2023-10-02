import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.const import MAIN_DIR, MONTH_LIST


class BypassRepository:
    
    
    def get_department_by_code(self, code: str):
        if code in ("03", ):
            return "Дагестанские огни"
        if code in ("07", ):
            return "Кизилюрт"
        if code in ("08", ):
            return "Кизляр"
        if code in ("01", ):
            return "Махачкала"
        if code in ("53", ):
            return "Унцукль"
    
    
    def serializeCommerce(self):
        path = os.path.join(MAIN_DIR, "Шаблоны расчетных ведомостей", "РВ Коммерческих потребителей")
        data = {}
        for book in os.listdir(path):
            file = load_workbook(os.path.join(path, book))
            print(os.path.join(path, book))
            sheet = file.worksheets[0]
            data = data | {
                sheet.cell(row, 3).value: [
                    1,
                    sheet.cell(row, 6).value,
                    f"{sheet.cell(row, 14).value}, {sheet.cell(row, 15).value}, {sheet.cell(row, 16).value}",
                    sheet.cell(row, 8).value,
                    "Юридическое лицо",
                    "",
                    self.get_department_by_code(sheet.cell(row, 3).value[3:5]),
                    sheet.cell(row, 19).value,
                    sheet.cell(row, 18).value,
                    "Исправен/Доступен",
                    sheet.cell(row, 21).value,
                    "",
                    "",
                    "",
                    "",
                    sheet.cell(row, 37).value,
                    sheet.cell(row, 35).value,
                    sheet.cell(row, 36).value,
                    sheet.cell(row, 9).value,
                    sheet.cell(row, 10).value,
                    sheet.cell(row, 11).value,
                ]
                for row in range(6, sheet.max_row + 1)
            }
        return data
    
    def get_last_three_month(self, month: str) -> list[str]:
        end_index = MONTH_LIST.index(month)
        start_index = end_index - 3 if end_index > 3 else 0
        return MONTH_LIST[start_index : end_index]
    
    def get_bypass_id(self, month: str) -> list[str]:
        data: dict[str, list] = {}
        path = os.path.join(
            MAIN_DIR,
            "Сводный баланс энергопотребления",
            f"Сводный баланс {datetime.now().year}",
            "Сводная ведомость потребителей",
            "Сводная ведомость Бытовых потребителей.xlsx"
        )
        book = load_workbook(path)
        for sheet_name in self.get_last_three_month(month):
            sheet = book[sheet_name]
            for row in range(6, sheet.max_row + 1):
                identifier = sheet.cell(row, 3).value
                if identifier not in data:
                    data[identifier] = []
                if sheet.cell(row, 31).value == "Обход":
                    data[identifier].append("Обход")
        return [key for key, value in data.items() if len(value) >= 3]
                
    def serializeIndividual(self, bypass_ids: list[str]):
        path = os.path.join(MAIN_DIR, "Шаблоны расчетных ведомостей", "РВ Бытовых потребителей")
        for book in os.listdir(path):
            file = load_workbook(os.path.join(path, book))
            sheet = file.worksheets[0]
            return {
                sheet.cell(row, 3).value: [
                    1,
                    sheet.cell(row, 6).value,
                    f"{sheet.cell(row, 14).value}, {sheet.cell(row, 15).value}, {sheet.cell(row, 16).value}",
                    sheet.cell(row, 8).value,
                    "Физическое лицо",
                    "",
                    self.get_department_by_code(sheet.cell(row, 3).value[3:5]),
                    sheet.cell(row, 19).value,
                    sheet.cell(row, 18).value,
                    "Исправен/Доступен",
                    sheet.cell(row, 21).value,
                    "",
                    "",
                    "",
                    "",
                    sheet.cell(row, 37).value,
                    sheet.cell(row, 35).value,
                    sheet.cell(row, 36).value,
                    sheet.cell(row, 9).value,
                    sheet.cell(row, 10).value,
                    sheet.cell(row, 11).value,
                ]
                for row in range(6, sheet.max_row + 1)
                if sheet.cell(row, 3).value in bypass_ids
            }
    
    
    