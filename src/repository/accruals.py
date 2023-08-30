import os
from datetime import datetime
from typing import Dict, Union
from abc import ABC, abstractmethod, abstractproperty

from openpyxl import load_workbook


MONTH_LIST = [
    "Январь", "Февраль", "Март",
    "Апрель", "Май", "Июнь",
    "Июль", "Август", 'Сентябрь',
    'Октябрь', 'Ноябрь', 'Декабрь'
]


class AccrualsRepository(ABC):

    @abstractmethod
    def getEmptyID():
        raise NotImplementedError

    @abstractmethod
    def getExpensesSvod():
        raise NotImplementedError

    @abstractmethod
    def getExpensesBicu():
        raise NotImplementedError
    
    @abstractproperty
    def getPath():
        path = ...
        return path

class AccrualsCommerceRepository(AccrualsRepository):
    """
    Класс для начисления показаний для коммерческих потребителей
    """
    

    def getExpensesBicu(self, month: str) -> Dict[str, Dict[str, int]]:
        path = "..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость БИКУ\\Сводная ведомость БИКУ.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(8, sheet.max_row + 1):
            index = sheet.cell(row, 3).value
            if index[6:8] == "MC":
                data[sheet.cell(row, 3).value] = {
                    "value": sheet.cell(row, 18).value
                }
        return data


    def getEmptyID(self, month) -> Dict[str, Dict[str, Union[str, int]]]:

        path = "..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            if not sheet.cell(row, 23).value:
                data[sheet.cell(row, 3).value] = {
                    "value": 0,
                    "index": sheet.cell(row, 23).value if len(sheet.cell(row, 39).value) > 6 and  sheet.cell(row, 39).value[6:8] == "MC" else None,
                    "method": None
                }
        return data
    
    def getExpensesSvod(self, month: str) ->  Dict[str, Dict[str, int]]:
        year = datetime.now().year
        path = f"..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            if sheet.cell(row, 21).value:
                data[sheet.cell(row, 3).value] = {
                    "value": int(sheet.cell(row, 22).value) - int(sheet.cell(row, 21).value if sheet.cell(row, 21).value else 0)
                }
        return data
    


class AccrualsInividualRepository(AccrualsRepository):
    
    def getEmptyID(self, month):
        path = "..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость Бытовых потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            if not sheet.cell(row, 23).value:
                data[sheet.cell(row, 3).value] = {
                    "value": [],
                    "index": sheet.cell(row, 23).value if len(sheet.cell(row, 39).value) > 6 and  sheet.cell(row, 39).value[6:8] == "MC" else None,
                    "method": None
                }
        return data
    
    def getExpensesSvod(self, month: str):
        year = datetime.now().year
        path = f"..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            if sheet.cell(row, 21).value:
                data[sheet.cell(row, 3).value] = {
                    "value": int(sheet.cell(row, 22).value) - int(sheet.cell(row, 21).value if sheet.cell(row, 21).value else 0)
                }
        return data


class AccrualsService:
    pass


class AccrualsCommerceService(AccrualsService):
    
    repo = AccrualsCommerceRepository()
    
    def __init__(self, month: str) -> None:
        self.month = month
        self.data = self.repo.getEmptyID(self.month)
    
    
    def checkPreviousYear(self):
        year = datetime.now().year - 1
        path = f"..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        data = self.repo.getExpensesSvod(self.month)

        if not os.path.exists(path):
            return

        for key in self.data.keys():
            if key in data.keys() and not self.data[key]["value"]:
                self.data[key]["value"] = data[key]["value"]
                self.data["method"] = "Начисление за пред. период"

    def checkControlBalance(self):
        data = self.repo.getExpensesBicu(self.month)
        for key in self.data.keys():
            if key in data.keys() and not self.data[key]["value"]:
                self.data[key]["value"] = data[key]["value"]
                
                self.data[key]["method"] = "Показания контрольного прибора учета"
    
    def checkNearestMonths(self):
        index =  MONTH_LIST.index(self.month)
        
        for month in reversed(MONTH_LIST[ :index]):
            data = self.repo.getExpensesSvod(month)
            for key in self.data.keys():
                if key in data.keys() and not self.data[key]["value"]:
                    self.data[key]["value"] = data[key]["value"]
                    self.data["method"] = "Начисление за пред. период"
    
    def makeAccruals(self):
        self.checkPreviousYear()
        self.checkNearestMonths()
        self.checkControlBalance()
    
    
    def insertAccruals(self):
        year = datetime.now().year
        path = path = f"..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path)
        sheet = file[self.month]
        
        self.makeAccruals()
        
        for row in range(6, sheet.max_row + 1):
            index = sheet.cell(row, 3).value
            if index in self.data.keys():
                 sheet.cell(row, 22).value = sheet.cell(row, 21) + self.data[index]["value"]
                 sheet.cell(row, 31).value = self.data[index]["method"]
        
        file.save(path)
            
        

class AccrualsIndividualService(AccrualsService):
    pass
