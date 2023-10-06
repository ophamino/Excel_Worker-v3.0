import os
import xlrd
from openpyxl import load_workbook



def read_excel_file(file_path):
    workbook = xlrd.open_workbook(file_path)

    data = {}

    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        cell_o5 = sheet.cell_value(4, 14)  # O5
        cell_f13 = sheet.cell_value(12, 5)  # F13
        data[cell_o5] = cell_f13
    return data


def insert_values(month: str):
    data = read_excel_file(os.path.join("..\\Dagenergy", "Шаблоны расчетных ведомостей", "АСКУЭ", "АСКУЭ.xls"))
    if data:
        values = list(data.keys())
        path = os.path.join("..\\Dagenergy", "Сводный баланс энергопотребления", "Сводный баланс 2023", "Сводная ведомость потребителей", "Сводная ведомость Коммерческих потребителей.xlsx")
        book = load_workbook(path, data_only=True)
        sheet = book[month]
        for row in range(6, sheet.max_row +1):
            divice = str(sheet.cell(row, 19).value)
            if divice in values:
                sheet.cell(row, 22).value = data[divice]
                data.pop(str(sheet.cell(row, 19).value))
        book.save(path)
    
    if data:
        path = os.path.join("..\\Dagenergy", "Сводный баланс энергопотребления", "Сводный баланс 2023", "Сводная ведомость потребителей", "Сводная ведомость Бытовых потребителей.xlsx")
        book = load_workbook(path, data_only=True)
        sheet = book[month]
        for row in range(6, sheet.max_row +1):
            if str(sheet.cell(row, 19).value) in values:
                sheet.cell(row, 22).value = data[str(sheet.cell(row, 19).value)]
        book.save(path)