import os

from datetime import datetime

from openpyxl import load_workbook

from src.utils.const import MONTH_LIST, MAIN_DIR
from src.repository.accruals import AccrualsCommerceRepository, AccrualsInividualRepository


class AccrualsService:
    pass


class AccrualsCommerceService(AccrualsService):
    
    repo = AccrualsCommerceRepository()
    
    def __init__(self, month: str) -> None:
        self.month = month
        self.data = self.repo.getEmptyID(self.month)
    
    
    def checkPreviousYear(self):
        year = datetime.now().year - 1
        path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        data = self.repo.getExpensesSvod(self.month)

        if not os.path.exists(path):
            return

        for key in self.data.keys():
            if key in data.keys() and not self.data[key]["value"]:
                self.data[key]["value"] = data[key]["value"]
                self.data["method"] = "Начисление за пред. период"

    def checkControlBalance(self):
        data = self.repo.getExpensesBicu(self.month)
        for key in self.data.keys():
            if key in data.keys() and not self.data[key]["value"]:
                self.data[key]["value"] = data[key]["value"]
                
                self.data[key]["method"] = "Показания контрольного прибора учета"
    
    def checkNearestMonths(self):
        index =  MONTH_LIST.index(self.month)
        
        for month in reversed(MONTH_LIST[ :index]):
            data = self.repo.getExpensesSvod(month)
            for key in self.data.keys():
                if key in data.keys() and not self.data[key]["value"]:
                    self.data[key]["value"] = data[key]["value"]
                    self.data["method"] = "Начисление за пред. период"
    
    def makeAccruals(self):
        self.checkPreviousYear()
        self.checkNearestMonths()
        self.checkControlBalance()
    
    
    def insertAccruals(self):
        year = datetime.now().year
        path = path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path)
        sheet = file[self.month]
        
        self.makeAccruals()
        
        for row in range(6, sheet.max_row + 1):
            index = sheet.cell(row, 3).value
            if index in self.data.keys():
                 sheet.cell(row, 22).value = sheet.cell(row, 21) + self.data[index]["value"]
                 sheet.cell(row, 31).value = self.data[index]["method"]
        
        file.save(path)
            
        

class AccrualsIndividualService(AccrualsService):
    repo = AccrualsInividualRepository()
    
    @staticmethod
    def calculate_average(expenses: list[int]):
        if len(expenses) == 6:
            return sum(expenses) / 6

    @staticmethod
    def calculate_standart(identifier: str) -> int:
        if identifier[3:7] == "0301":
            return 256
        if identifier[3:7] in ("0101", "0701"):
            return 301
        return 0
    
    
    def serialize_accruals(self, month: str):
        empty_data = self.repo.getEmptyID(month)
        average = self.repo.get_total_expenses(month)
        for month in average:
            for identifier in average[month]:
                if identifier in empty_data:
                    empty_data["values"].append(average[month][identifier]["expenses"])
        return empty_data
    
    def insert_values(self, month: str) -> None:
        data = self.serialize_accruals(month)
        path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость Бытовых потребителей.xlsx"
        book = load_workbook(path, data_only=True)
        sheet = book[month]
        for row in range(6, sheet.max_row +1):
            identifier = sheet.cell(row, 3).value
            if identifier in data:
                if len(data[identifier]["values"]) == 6:
                    sheet.cell(row, 22).value = sheet.cell(row, 21).value + self.calculate_average(data[identifier]["values"])
                    sheet.cell(row, 31).value == "Среднемесячное"
                else:
                    sheet.cell(row, 22).value = sheet.cell(row, 21).value + self.calculate_standart(identifier)
                    sheet.cell(row, 31).value == "Нормативное начисление"
        book.save(path)
            
    
    