import os
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from service.base import BaseService
from repository.bicu import BicuReposiitory
from utils.const import MONTH_LIST, DEPARTAMENT
from utils.excel import open_excel
from utils.base import clean_directory, resource_path


class BicuService(BaseService):
    """Класс, отвечающий за работу сервиса БИКУ"""
    repository = BicuReposiitory()
    
    def formating_bicu(self, sheet: Worksheet) -> None:
        """
        Функция для форматирования файлов 'БИКУ': Нумерация и вставка формул

        Args:
            sheet (Worksheet): Лист Excel, который нужно отформатирвать
        """
        for row in range(8, sheet.max_row + 1):
            sheet.cell(row=row, column=1).value = row - 7
            sheet["R{0}".format(row)] = "=Q{0}-P{0}".format(row)
            sheet["S{0}".format(row)] = "=R{0}*O{0}".format(row)
            sheet["W{0}".format(row)] = "=S{0}+T{0}+U{0}+V{0}".format(row)
            
        letter_list = ["R", "S", "T", "U", "V", "W"]
            
        for letter in letter_list:
            sheet[f"{letter}4"] = f'=SUMIFS({letter}8:{letter}{sheet.max_row},$E8:$E{sheet.max_row},"Прием электроэнергии")'
            sheet[f"{letter}5"] = f'=SUMIFS({letter}8:{letter}{sheet.max_row},$E8:$E{sheet.max_row},"Передача электроэнергии")'
            sheet[f"{letter}6"] = f'={letter}4-{letter}5'
    
    def get_svod_if_exists(self, path: str) -> Workbook:
        """
        Функция загружает сводную ведомость, если не существует - создает

        Args:
            path (str): Путь к файлу

        Returns:
            Workbook: Загруженная книга
        """
        if not os.path.exists(path):
            book = open_excel("template\Сводная ведомость БИКУ.xlsx")
            book.save(path)

        return open_excel(path)

    def collect_data_in_svod(self, month: int) -> None:
        """
        Функия для объединения расчетных ведомостей в одну сводную ведомости за определеннный месяц

        Args:
            month (int): Номер месяца
        """
        month = MONTH_LIST[month - 1]
        data = self.repository.collect_svod(month)
        
        path = (f"{self.directory}\Сводный баланс энергопотребления\Сводный баланс 2023\Сводная ведомость БИКУ\Сводная ведомость БИКУ.xlsx")
        book = self.get_svod_if_exists(path)
        sheet = book[month]
        
        self.insert_data(sheet, data)
        self.formating_bicu(sheet)
        book.save(path)
        
        logging.info(f"Сформирована сводная ведомость БИКУ за {month}")
    
    def create_new_statement(self, month: int) -> None:
        """
        Функия для формирования расчтный ведомостей БИКУ

        Args:
            month (int): Месяц за который нужно сформирвоать расчетную ведомость
        """
        month = MONTH_LIST[month - 1]
        data = self.repository.get_new_statement(month)
        path = resource_path(f"{self.directory}\Шаблоны расчетных ведомостей\РВ БИКУ")
        clean_directory(path)
        
        for departament_id, name in DEPARTAMENT.items():
            book = open_excel("template/Расчетная ведомость БИКУ.xlsx")
            sheet = book.worksheets[0]
            
            self.insert_data_by_departament(sheet, data, departament_id)
            self.formating_bicu(sheet)
            book.save(f"{path}\РВ БИКУ {name} {datetime.now().year} {month}.xlsx")
        
        logging.info(f"СФормированы Расчетные ведомости БИКУ за {month}")
