import os
from typing import Dict, Optional, List
from abc import ABC, abstractmethod

from openpyxl import load_workbook


class AbstractRepository(ABC):
    
    @abstractmethod
    def serialize_data(self, path: str):
        raise NotImplementedError
    
    @abstractmethod
    def serialize_all_data(self, path: str):
        raise NotImplementedError


class ExcelRepository(AbstractRepository):
    skip_rows: Optional[int] = None
    id_col: Optional[int]  = None
    keys: Optional[List[str]] = None

    def serialize_data(self, path: str) -> Dict[str,  Dict]:
        """
        Функция для сериализации данных определнного файла

        Args:
            path (str): Путь к файлу

        Returns:
            Dict[str, Dict[str, Any]]: Итоговые данные
        """
        data = dict()
        file = load_workbook(path)
        sheet = file.worksheets[0]
        for row in range(self.skip_rows, sheet.max_row + 1):
            data[sheet.cell(row, self.id_col).value] = {
                key: sheet.cell(row, col).value for col, key in enumerate(self.keys, 1)
            }
        return data
    
    def serialize_all_data(self, path: str) -> Dict[str, Dict]:
        """
        Функия для сериализации данных всех файлов с определенной директории

        Args:
            path (str): Путь к директории

        Returns:
            Dict[str, Dict[str, Any]]: Итоговые данные
        """
        files, data = os.listdir(path), dict()
        for file in files:
            data = data | self.serialize_data(f"{path}\{file}")
        return data
