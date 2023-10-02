from datetime import datetime
from typing import Dict, Union
from abc import ABC, abstractmethod

from openpyxl import load_workbook, Workbook

from src.utils.const import MAIN_DIR, MONTH_LIST


class AccrualsRepository(ABC):

    @abstractmethod
    def getEmptyID():
        raise NotImplementedError

    @abstractmethod
    def getExpensesSvod():
        raise NotImplementedError


class AccrualsCommerceRepository(AccrualsRepository):
    """
    Класс для начисления показаний для коммерческих потребителей
    """
    

    def getExpensesBicu(self, month: str) -> Dict[str, Dict[str, int]]:
        path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость БИКУ\\Сводная ведомость БИКУ.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(8, sheet.max_row + 1):
            index = sheet.cell(row, 3).value
            if index[6:8] == "MC":
                data[sheet.cell(row, 3).value] = {
                    "value": sheet.cell(row, 18).value
                }
        return data


    def getEmptyID(self, month) -> Dict[str, Dict[str, Union[str, int]]]:

        path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            if not sheet.cell(row, 23).value:
                data[sheet.cell(row, 3).value] = {
                    "value": 0,
                    "index": sheet.cell(row, 23).value if len(sheet.cell(row, 39).value) > 6 and  sheet.cell(row, 39).value[6:8] == "MC" else None,
                    "method": None
                }
        return data
    
    def getExpensesSvod(self, month: str) ->  Dict[str, Dict[str, int]]:
        year = datetime.now().year
        path = f"..\\Dagenergy\\Сводный баланс энергопотребления\\Сводный баланс {year}\\Сводная ведомость потребителей\\Сводная ведомость Коммерческих потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[month]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            if sheet.cell(row, 21).value:
                data[sheet.cell(row, 3).value] = {
                    "value": int(sheet.cell(row, 22).value) - int(sheet.cell(row, 21).value if sheet.cell(row, 21).value else 0)
                }
        return data
    


class AccrualsInividualRepository(AccrualsRepository):
    
    def getEmptyID(self, month: str) -> Dict[str, Dict[str, int]]:
        path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость Бытовых потребителей.xlsx"
        file = load_workbook(path)
        sheet = file[month]
        return {
            sheet.cell(row, 3).value : {
                "values": [],
                "method": None,
                "expenses": 0,
            }
            for row in range(6, sheet.max_row + 1)
            if not sheet.cell(row, 22).value
        }
        
        
    def getExpensesSvod(self, month: str, book: Workbook) -> Dict[str, Dict[str, int]]:
        sheet = book[month]
        return {
            sheet.cell(row, 3).value: {
                "expenses": sheet.cell(row, 22).value - sheet.cell(row, 21).value
            }
            for row in range(6, sheet.max_row + 1)
        }
    
    def get_total_expenses(self, month: str):
        path = f"{MAIN_DIR}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость Бытовых потребителей.xlsx"
        book = load_workbook(path, data_only=True)
        month_index = MONTH_LIST.index(month)
        if month_index >= 5:
            return {
                month: self.getExpensesSvod(month, book)
                for month in MONTH_LIST[month_index - 5: month_index]
            }
