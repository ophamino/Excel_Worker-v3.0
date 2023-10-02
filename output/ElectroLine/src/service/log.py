import os
from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl import load_workbook
from src.utils.const import MAIN_DIR
from src.utils.base import resource_path


class Log:
    
    change_path = f"{MAIN_DIR}\\Реестровая база данных\\Реестр потребителей\\Реестр потребителей.xlsx"
    static_path = resource_path(".\\src\\template\\consumers.xlsx")
    upload_to = f"{MAIN_DIR}\\Реестровая база данных\Реестр потребителей\Журнал изменений.xlsx"
    col_id = 4
    
    def serialaze_changes(self)-> dict[str, list[Cell]]:
        data = {}
        path = self.change_path
        file = load_workbook(path, data_only=True)
        sheet = file.worksheets[0]
        
        for row in range(9, sheet.max_row + 1):
            data[sheet.cell(row=row, column=4).value] = {
                "values": [sheet.cell(row=row, column=col).value for col in range(1, 55)],
                "coordinate": [sheet.cell(row=row, column=col).coordinate for col in range(1, 55)]
            }
        return data
    
    def serialaze_static(self) -> dict[str, list[Cell]]:
        data = {}
        path =  self.static_path
        file = load_workbook(path, data_only=True)
        sheet = file.worksheets[0]
        
        for row in range(9, sheet.max_row + 1):
            data[sheet.cell(row=row, column=self.col_id).value] = {
                "values": [sheet.cell(row=row, column=col).value for col in range(1, 55)],
                "coordinate": [sheet.cell(row=row, column=col).coordinate for col in range(1, 55)]
            }
        
        return data
    
    def collect_changes(self) -> list:
        data = []
        changes = self.serialaze_changes()
        static = self.serialaze_static()
        
        for key in changes.keys():
            if key not in static.keys():
                data.append([1, key, "", "", "", "", "Добавлено", datetime.now().date(), datetime.now().time()])
            
            if key in static.keys():
                new_values = static[key]["values"]
                old_values = changes[key]["values"]
                if new_values != old_values:
                    for index in range(len((new_values))):
                        if new_values[index] != old_values[index]:
                            coordinate = changes[key]["coordinate"][index]
                            data.append([1, key, "", coordinate, new_values[index], old_values[index], "Изменено", datetime.now().date(), datetime.now().time()])
                static.pop(key)
            
        if static.keys():
            for key in static.keys():
                data.append([1, key, "", "", "", "", "Удалено", datetime.now().date(), datetime.now().time()])
        return data
    
    def delete_nulls(self, sheet: Worksheet):
        sheet.delete_rows(4, 10)
        
    
    def insert_changes(self):
        data = self.collect_changes()
        path = path = self.upload_to
        if not os.path.exists(path):
            file = load_workbook(resource_path('.\\src\\template\\log.xlsx'))
            file.save(path)
        file = load_workbook(path)
        sheet = file[str(datetime.now().year)]
        
        for row in data:
            if (not row[4]) and (not row[5]) and (row[6] == "Изменено"):
                continue
            sheet.append(row)
        
        file.save(path)
    
    def search_changes(self):
        self.insert_changes()
        file = load_workbook(self.change_path, data_only=True)
        file.save(self.static_path)


class BicuLog(Log):
    changs_path = f"{MAIN_DIR}\\Реестровая база данных\\Реестр БИКУ\\Реестр БИКУ.xlsx"
    col_id = 3
    static_path = resource_path(".\\src\\template\\bicu.xlsx")
    upload_to = f"{MAIN_DIR}\\Реестровая база данных\\Реестр БИКУ\\Журнал изменений.xlsx"
