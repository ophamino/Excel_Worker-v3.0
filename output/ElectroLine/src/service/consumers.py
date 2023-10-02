import os
import logging
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.const import MONTH_LIST, DEPARTAMENT
from src.utils.base import clean_directory, resource_path
from src.service.base import BaseService
from src.repository.consumers import ComsumerRepository

class ConsumersService(BaseService):
    repository = ComsumerRepository()
    
    
    def formating_consumers(self, sheet: Worksheet):
        """
        Функция для форматирования Сводной ведомости потребителей: нумерация и вставка формул

        Args:
            sheet (Worksheet): Лист Excel, который нужно отформатирвать
        """
        for row in range(6, sheet.max_row +1):
            sheet.cell(row, 1).value = row - 5
            sheet.cell(row, 23).value = "=IF(V{0},V{0}-U{0},0)".format(row)
            sheet.cell(row, 24).value = "=W{0}*T{0}".format(row)
            sheet.cell(row, 25).value = "=IF(ISBLANK($AL${0}),0,ROUND(($X${0}*$AL${0}),0))".format(row)
            sheet.cell(row, 29).value = "=X{0}+Y{0}+Z{0}+AA{0}+AB{0}".format(row)
        sheet.cell(4, 23).value = "=SUM(W6:W{0})".format(sheet.max_row)
        sheet.cell(4, 24).value = "=SUM(X6:X{0})".format(sheet.max_row)
        sheet.cell(4, 25).value = "=SUM(Y6:Y{0})".format(sheet.max_row)
        sheet.cell(4, 26).value = "=SUM(Z6:Z{0})".format(sheet.max_row)
        sheet.cell(4, 27).value = "=SUM(AA6:AA{0})".format(sheet.max_row)
        sheet.cell(4, 28).value = "=SUM(AB6:AB{0})".format(sheet.max_row)
        sheet.cell(4, 29).value = "=SUM(AC6:AC{0})".format(sheet.max_row)
    
    def get_svod_if_exists(self, path: str) -> Workbook:
        """
        Функция загружает сводную ведомость, если не существует - создает

        Args:
            path (str): Путь к файлу

        Returns:
            Workbook: Загруженная книга
        """
        if not os.path.exists(path):
            book = load_workbook("src\\template\\svod_consumers.xlsx")
            book.save(path)

        return load_workbook(path)
    
    def collect_commerce_svod(self, month: int) -> None:
        """
        Функция для формирования сводной ведомости коммерческих потребителей

        Args:
            month (int): Номер месяца, за который нужно сформировать сводную ведомость
        """
        path = f"{self.directory}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей"
        data = self.repository.serialize_all_data(f"{path}\\РВ Потребителей {month}\\РВ Коммерческих потребителей")
        book = self.get_svod_if_exists(f"{path}\\Сводная ведомость Коммерческих потребителей.xlsx")
        sheet = book[month]

        self.insert_data(sheet, data)
        self.formating_consumers(sheet)
        book.save(f"{path}\\Сводная ведомость Коммерческих потребителей.xlsx")

        logging.info(f"Сформирована сводная ведомость коммерческих потребителей за {month}")

        
    def collect_individual_svod(self, month: str) -> None:
        """
        Функция для формирования сводной ведомости бытовых потребителей

        Args:
            month (str): Название месяца
        """
        path = f"{self.directory}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей"

        data = self.repository.serialize_all_data(f"{path}\\РВ Потребителей {month}\\РВ Бытовых потребителей")
        book = self.get_svod_if_exists(f"{path}\\Сводная ведомость Бытовых потребителей.xlsx")
        sheet = book[month]

        self.insert_data(sheet, data)
        self.formating_consumers(sheet)
        book.save(f"{path}\\Сводная ведомость Бытовых потребителей.xlsx")

        logging.info(f"Сформирована сводная ведомость бытовых потребителей за {month}")
        
    def collect_total_svod(self, month: int) -> None:
        data = self.repository.get_total_svod_data(month)
        
        path = f"{self.directory}\\Сводный баланс энергопотребления\\Сводный баланс 2023\\Сводная ведомость потребителей\\Сводная ведомость потребителей.xlsx"
        book = self.get_svod_if_exists(path)
        sheet = book[month]
        
        self.insert_data(sheet, data)
        self.formating_consumers(sheet)
        
        book.save(path)
        logging.info(f"Сформирована сводная ведомость потребителей за {month}")
    
    def create_individual_statement(self, month: int):
        data = self.repository.get_statment(month, "ch")
        path = resource_path(f"{self.directory}\\Шаблоны расчетных ведомостей\\РВ Бытовых потребителей")
        clean_directory(path)
        
        for departament_id, name in DEPARTAMENT.items():
            book = load_workbook("src\\template\\rv_consumers.xlsx")
            sheet = book.worksheets[0]
            
            self.insert_data_by_departament(sheet, data, departament_id)
            self.formating_consumers(sheet)
            book.save(f"{path}\\РВ Бытовых Потребителей {name} {datetime.now().year} {month}.xlsx")
        logging.info(f"СФормированы Расчетные ведомости Бытовых потребителей за {month}")
            
    def create_commerce_statement(self, month: int):
        data = self.repository.get_statment(month, "cc")
        path = resource_path(f"{self.directory}\\Шаблоны расчетных ведомостей\\РВ Коммерческих потребителей")
        clean_directory(path)
        
        for departament_id, name in DEPARTAMENT.items():
            book = load_workbook("src\\template\\rv_consumers.xlsx")
            sheet = book.worksheets[0]
            
            self.insert_data_by_departament(sheet, data, departament_id)
            self.formating_consumers(sheet)
            book.save(f"{path}\\РВ Коммерческих потребителей {name} {datetime.now().year} {month}.xlsx")
        logging.info(f"СФормированы Расчетные ведомости Коммерческих потребителей за {month}")
